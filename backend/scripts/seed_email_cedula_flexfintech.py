"""
Seed historico email->cedula para FlexFintech desde los Excel operativos.

Lee varias hojas que tienen (cedula, email, nombre) y UPSERT a
`historico_email_cedula` para el tenant FlexFintech. Idempotente.

Hojas leídas:
- Consolidado Reclamos.xlsx:
    * Mails              (cols: DOCUMENTO, NOMBRE, MAIL, FECHA DE CARGA)
    * Reclamos           (cols: DNI/CEDULA CLIENTE, APELLIDO Y NOMBRE,
                                DATO CONTACTO, FECHA RECLAMO)
    * CONSOLIDADO SANTANDER (cols: DOCUMENTO, NOMBRE, MAIL, FECHA DE INGRESO)
    * CONSOLIDADO BOGOTA    (idem)
- Consolidado Defcon.xlsx:
    * Colombia           (cols: Cedula Cliente, Nombre Y Apellido, Email,
                                Fecha Ingreso)

Uso (dentro del container backend):
    python -m scripts.seed_email_cedula_flexfintech \\
        --reclamos /tmp/reclamos.xlsx --defcon /tmp/defcon.xlsx
    python -m scripts.seed_email_cedula_flexfintech --dry-run [...]
"""

from __future__ import annotations

import argparse
import asyncio
import logging
import os
import re
import sys
from datetime import datetime
from typing import Iterable

import asyncpg
import openpyxl

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("seed_email_cedula_ff")

TENANT_FF = "f7e8d9c0-b1a2-3456-7890-123456abcdef"

# Email válido: minimal pero sólido. Excluye placeholders típicos.
_EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")
_PLACEHOLDER_EMAILS = {"n/a", "na", "ninguno", "-", "sin email", "no tiene"}


def _clean(s) -> str:
    if s is None:
        return ""
    return str(s).replace("\xa0", " ").strip()


def _normalize_cedula(s) -> str:
    """Solo dígitos. '10.000.000' → '10000000'. Rechaza si <6 dígitos."""
    digits = re.sub(r"\D", "", str(s) if s is not None else "")
    return digits if len(digits) >= 6 else ""


def _normalize_email(s) -> str:
    e = _clean(s).lower()
    if not e or e in _PLACEHOLDER_EMAILS:
        return ""
    return e if _EMAIL_RE.match(e) else ""


def _parse_date(s) -> datetime | None:
    if s is None or s == "":
        return None
    if isinstance(s, datetime):
        return s
    try:
        return datetime.strptime(str(s)[:10], "%Y-%m-%d")
    except Exception:
        return None


# ──────────────────────────────────────────────────────────────────────────
# Lectores por hoja — devuelven iterable de (cedula, email, nombre, fecha)
# ──────────────────────────────────────────────────────────────────────────

def _read_sheet(
    xlsx_path: str,
    sheet: str,
    col_cedula: int,
    col_nombre: int,
    col_email: int,
    col_fecha: int,
    skip_header: int = 1,
) -> Iterable[tuple[str, str, str, datetime | None]]:
    """Generic reader por índices de columna (0-based)."""
    if not os.path.exists(xlsx_path):
        logger.warning("  ⚠️  %s no encontrado — skip hoja '%s'", xlsx_path, sheet)
        return
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        logger.warning("  ⚠️  hoja '%s' no existe en %s — skip", sheet, xlsx_path)
        return
    ws = wb[sheet]
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= skip_header:
            continue
        if col_cedula >= len(row) or col_email >= len(row):
            continue
        cedula = _normalize_cedula(row[col_cedula])
        email = _normalize_email(row[col_email])
        if not cedula or not email:
            continue
        nombre = _clean(row[col_nombre] if col_nombre < len(row) else "")
        fecha = _parse_date(row[col_fecha] if col_fecha < len(row) else None)
        yield (cedula, email, nombre, fecha)


def cargar_mapeo(reclamos: str, defcon: str) -> list[dict]:
    """Recopila pares de todas las hojas, deduplica por email (el primero gana)."""
    visto = {}  # email -> dict
    fuentes_count = {}

    sources = [
        # (file, sheet, cedula_col, nombre_col, email_col, fecha_col, fuente)
        (reclamos, "Mails",              2, 4, 5, 7, "rec:Mails"),
        (reclamos, "Reclamos",           5, 6, 14, 7, "rec:Reclamos"),
        (reclamos, "CONSOLIDADO SANTANDER", 2, 3, 4, 5, "rec:CONS-SANTANDER"),
        (reclamos, "CONSOLIDADO BOGOTA",    2, 3, 4, 5, "rec:CONS-BOGOTA"),
        (defcon,   "Colombia",           2, 5, 15, 7, "def:Colombia"),
    ]

    for path, sheet, cc, cn, ce, cf, fuente in sources:
        cont = 0
        for cedula, email, nombre, fecha in _read_sheet(path, sheet, cc, cn, ce, cf):
            cont += 1
            if email not in visto:
                visto[email] = {
                    "email": email,
                    "cedula": cedula,
                    "nombre": nombre,
                    "fecha": fecha,
                    "fuente": fuente,
                }
        fuentes_count[fuente] = cont
        logger.info("  hoja '%s': %d pares válidos extraídos", sheet, cont)

    logger.info("Resumen por fuente: %s", fuentes_count)
    logger.info("Pares únicos (deduplicados por email): %d", len(visto))
    return list(visto.values())


async def upsert(conn: asyncpg.Connection, docs: list[dict], dry_run: bool) -> dict:
    cont = {"insertados": 0, "actualizados": 0}
    if dry_run:
        for d in docs[:10]:
            logger.info("  DRY [%s] %s -> %s (%s)",
                        d["fuente"], d["email"], d["cedula"], (d["nombre"] or "")[:30])
        if len(docs) > 10:
            logger.info("  ... y %d más", len(docs) - 10)
        return cont

    for d in docs:
        # ON CONFLICT (cliente_id, email) DO UPDATE
        existing = await conn.fetchrow(
            "SELECT id, primera_vez FROM historico_email_cedula "
            "WHERE cliente_id = $1::uuid AND email = $2",
            TENANT_FF, d["email"],
        )
        if existing:
            await conn.execute(
                """UPDATE historico_email_cedula
                   SET cedula = $1, nombre = COALESCE(NULLIF($2,''), nombre),
                       ultima_vez = CURRENT_TIMESTAMP, fuente = $3
                   WHERE id = $4""",
                d["cedula"], d["nombre"], d["fuente"], existing["id"],
            )
            cont["actualizados"] += 1
        else:
            await conn.execute(
                """INSERT INTO historico_email_cedula
                      (cliente_id, email, cedula, nombre, primera_vez, ultima_vez, fuente)
                   VALUES ($1::uuid, $2, $3, NULLIF($4,''),
                           COALESCE($5, CURRENT_TIMESTAMP),
                           COALESCE($5, CURRENT_TIMESTAMP), $6)""",
                TENANT_FF, d["email"], d["cedula"], d["nombre"], d["fecha"], d["fuente"],
            )
            cont["insertados"] += 1
    return cont


async def main_async(args: argparse.Namespace) -> int:
    docs = cargar_mapeo(args.reclamos, args.defcon)
    if not docs:
        logger.warning("Sin pares válidos. Saliendo.")
        return 0

    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        logger.error("DATABASE_URL no configurada")
        return 2

    conn = await asyncpg.connect(db_url)
    try:
        await conn.execute("SET app.is_superuser = 'true'")
        cont = await upsert(conn, docs, args.dry_run)
        logger.info(
            "Hecho. insertados=%d actualizados=%d (dry_run=%s)",
            cont["insertados"], cont["actualizados"], args.dry_run,
        )
    finally:
        await conn.close()
    return 0


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--reclamos", required=True, help="path Consolidado Reclamos.xlsx")
    p.add_argument("--defcon", required=True, help="path Consolidado Defcon.xlsx")
    p.add_argument("--dry-run", action="store_true")
    return p.parse_args()


if __name__ == "__main__":
    sys.exit(asyncio.run(main_async(parse_args())))
