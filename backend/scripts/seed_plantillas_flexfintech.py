"""
Seed de plantillas FlexFintech extraídas de los Excel operativos.

Lee `Rtas` + `RTA DC` de `Consolidado Reclamos.xlsx` y UPSERT a
`plantillas_respuesta` para el tenant FlexFintech. Idempotente: re-correr
es seguro (UPSERT por (cliente_id, problematica)).

Uso (dentro del container backend):
    python -m scripts.seed_plantillas_flexfintech --xlsx /tmp/reclamos.xlsx
    python -m scripts.seed_plantillas_flexfintech --xlsx /tmp/reclamos.xlsx --dry-run
"""

from __future__ import annotations

import argparse
import asyncio
import logging
import os
import sys

import asyncpg
import openpyxl

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("seed_plantillas_ff")

TENANT_FF = "f7e8d9c0-b1a2-3456-7890-123456abcdef"

# ──────────────────────────────────────────────────────────────────────────
# Mapping (TIPO, MOTIVO) → (slug, tipo_workflow, contexto, keywords_trigger)
# - slug: identifica la plantilla (UPSERT key junto con cliente_id+workflow)
# - keywords_trigger: si NO vacía, `detectar_problematica` matchea esto y
#   dispara la plantilla automáticamente. Si vacía → admin la usa desde el
#   dashboard manualmente.
# ──────────────────────────────────────────────────────────────────────────

MAPPING_RTAS = {
    # ─── disparables automáticamente ─────────────────────────────────────
    ("PAZ Y SALVO", "PEDIDO PAZ Y SALVO"): (
        "PEDIDO_PAZ_Y_SALVO", "ATENCION_CLIENTE",
        "Cliente solicita certificado de paz y salvo / libre de deuda",
        ["paz y salvo", "certificado de cancelación", "certificado de cancelacion",
         "libre de deuda", "obligacion cancelada", "obligación cancelada"],
    ),
    ("COMPROBANTE", "COMPROBANTE RECIBIDO"): (
        "COMPROBANTE_RECIBIDO", "ATENCION_CLIENTE",
        "Cliente envía comprobante de pago — acuse de recibo con documento adjunto",
        ["adjunto comprobante", "adjunto el comprobante", "envío comprobante",
         "comprobante de pago", "comprobante de transferencia"],
    ),

    # ─── manuales (admin desde dashboard) ────────────────────────────────
    ("PAZ Y SALVO", "Manda comprobante del dia y pide PYS"): (
        "PAZ_Y_SALVO_COMPROBANTE_RECIENTE", "ATENCION_CLIENTE",
        "Cliente envía comprobante y pide PYS — pago aún en acreditación", [],
    ),
    ("PAZ Y SALVO BOGOTA", "Porque JJ está atrasado cn las remesas de bogota"): (
        "PAZ_Y_SALVO_BOGOTA_DEMORA_JJ", "ATENCION_CLIENTE",
        "PYS Bogotá demorado por remesas de la agencia JJ", [],
    ),
    ("IDENTIFICACION CLIENTE", "PEDIR DOCUMENTO"): (
        "PEDIR_DOCUMENTO_CLIENTE", "ATENCION_CLIENTE",
        "Solicitar número de documento al cliente para asistirlo", [],
    ),
    ("COMPROBANTE", "COMPROBANTE SIN DOCUMENTO"): (
        "COMPROBANTE_SIN_DOCUMENTO", "ATENCION_CLIENTE",
        "Comprobante recibido pero falta documento del cliente", [],
    ),
    # NOTA: "ANTENCION" (con N extra) NO es un typo del código — es como
    # literalmente aparece en el Excel del cliente. Las 3 keys siguientes
    # preservan ese typo a propósito para que el match exacto funcione.
    # Si en algún momento el Excel se "limpia" a "ATENCION", actualizar acá.
    ("ANTENCION AL CLIENTE", "CASO EN REVISION"): (
        "CASO_EN_REVISION", "ATENCION_CLIENTE",
        "Confirmar al cliente que su caso ya fue asignado a revisión", [],
    ),
    ("ANTENCION AL CLIENTE", "PENDIENTE RTA CLIENTE"): (
        "CASO_PENDIENTE_RTA_CLIENTE", "ATENCION_CLIENTE",
        "Recordatorio al cliente — necesitamos su respuesta para continuar", [],
    ),
    ("ANTENCION AL CLIENTE", "CIERRE NO RTA CLIENTE"): (
        "CASO_CIERRE_NO_RTA_CLIENTE", "ATENCION_CLIENTE",
        "Cierre del caso por falta de respuesta del cliente", [],
    ),
    # Datos de pago — uno por banco/país
    ("DATOS DE PAGO", "ARGENTINA FIDEI"): (
        "DATOS_PAGO_ARGENTINA_FIDEI", "ATENCION_CLIENTE",
        "Datos cuenta Santander Argentina para pagos al fideicomiso", [],
    ),
    ("DATOS DE PAGO", "BANCOLOMBIA"): (
        "DATOS_PAGO_BANCOLOMBIA", "ATENCION_CLIENTE",
        "Datos cuenta Bancolombia para pagos en Colombia", [],
    ),
    ("DATOS DE PAGO", "BANCO DE OCCIDENTE"): (
        "DATOS_PAGO_BANCO_OCCIDENTE", "ATENCION_CLIENTE",
        "Datos cuenta Banco de Occidente para pagos en Colombia", [],
    ),
    ("DATOS DE PAGO", "URUGUAY"): (
        "DATOS_PAGO_URUGUAY", "ATENCION_CLIENTE",
        "Datos cuenta Banco República para pagos en Uruguay", [],
    ),
    # Agencias — cada una específica
    ("AGENCIA", "GDF ARGENTINA"): (
        "AGENCIA_GDF_AR", "ATENCION_CLIENTE",
        "Caso derivado al estudio GDF Argentina", [],
    ),
    ("AGENCIA", "COACTIVA"): (
        "AGENCIA_COACTIVA", "ATENCION_CLIENTE",
        "Caso derivado a la agencia COACTIVA GROUP", [],
    ),
    ("AGENCIA", "RCI"): (
        "AGENCIA_RCI", "ATENCION_CLIENTE",
        "Caso derivado a la agencia RCI", [],
    ),
    ("AGENCIA", "GDF COLOMBIA"): (
        "AGENCIA_GDF_CO", "ATENCION_CLIENTE",
        "Caso derivado al estudio GDF Colombia", [],
    ),
    ("AGENCIA", "PUNTUALMENTE"): (
        "AGENCIA_PUNTUALMENTE", "ATENCION_CLIENTE",
        "Caso derivado a la agencia Puntualmente", [],
    ),
    ("AGENCIA", "JJ ABOGADOS"): (
        "AGENCIA_JJ_ABOGADOS", "ATENCION_CLIENTE",
        "Caso derivado a la agencia JJ Abogados", [],
    ),
    ("AGENCIA", "CYC"): (
        "AGENCIA_CYC_SERVICES", "ATENCION_CLIENTE",
        "Caso derivado a la agencia C&C Services", [],
    ),
    ("AGENCIA", "ABOGADOS RECOVERY"): (
        "AGENCIA_ABOGADOS_RECOVERY", "ATENCION_CLIENTE",
        "Caso derivado a la agencia Abogados Recovery", [],
    ),
    ("AGENCIA", "AG115 - JJ MEDELLIN"): (
        "AGENCIA_AG115_JJ_MEDELLIN", "ATENCION_CLIENTE",
        "Caso derivado a JJ Medellín (AG115)", [],
    ),
    ("AGENCIA", "AG116 - JJ BOGOTA"): (
        "AGENCIA_AG116_JJ_BOGOTA", "ATENCION_CLIENTE",
        "Caso derivado a JJ BPO Bogotá (AG116)", [],
    ),
    ("AGENCIA", "JJ - BOGOTÁ 2"): (
        "AGENCIA_JJ_BOGOTA_V2", "ATENCION_CLIENTE",
        "Caso derivado a JJ BPO Bogotá (variante 2)", [],
    ),
    ("AGENCIA", "122 - GG GROUP"): (
        "AGENCIA_GG_GROUP_AG122", "ATENCION_CLIENTE",
        "Caso derivado a GG GROUP (AG122)", [],
    ),
    ("AGENCIA", "121 - DELTA CREDIT"): (
        "AGENCIA_DELTA_CREDIT_AG121", "ATENCION_CLIENTE",
        "Caso derivado a Delta Credit (AG121)", [],
    ),
    ("AGENCIA", "CCALING"): (
        "AGENCIA_CALLING", "ATENCION_CLIENTE",
        "Caso derivado a la agencia Calling", [],
    ),
    ("AGENCIA", "VIRTUS - AG117"): (
        "AGENCIA_VIRTUS_AG117", "ATENCION_CLIENTE",
        "Caso derivado a Virtus (AG117)", [],
    ),
    ("AGENCIA", "122- GG GROUP"): (
        "AGENCIA_GG_GROUP_AG122_V2", "ATENCION_CLIENTE",
        "Caso derivado a GG GROUP (AG122 variante)", [],
    ),
    ("AGENCIA", "123 - GERENCIAR"): (
        "AGENCIA_GERENCIAR_AG123", "ATENCION_CLIENTE",
        "Caso derivado a la agencia Gerenciar (AG123)", [],
    ),
    # Otras
    ("ENTIDAD ORIGEN", "AT CLIENTE FLB"): (
        "AT_CLIENTE_FALABELLA", "ATENCION_CLIENTE",
        "Teléfono de atención al cliente de Falabella", [],
    ),
    ("DETALLE DEUDA", "FALABELLA"): (
        "DETALLE_DEUDA_FALABELLA", "ATENCION_CLIENTE",
        "Detalle de deuda Falabella con saldo a regularizar", [],
    ),
    ("CONVENIO", "RECORDATORIO CONVENIO FLB"): (
        "CONVENIO_RECORDATORIO_FALABELLA", "ATENCION_CLIENTE",
        "Recordatorio de cuota convenio Falabella", [],
    ),
    ("CONVENIO", "RECORDATORIO CONVENIO RAPPI"): (
        "CONVENIO_RECORDATORIO_RAPPI", "ATENCION_CLIENTE",
        "Recordatorio de cuota convenio Rappi Pay", [],
    ),
    ("CONVENIO", "RECORDATORIO CONVENIO COPPEL"): (
        "CONVENIO_RECORDATORIO_COPPEL", "ATENCION_CLIENTE",
        "Recordatorio de cuota convenio Coppel", [],
    ),
    ("CONVENIO", "CONVENIO VENCIDO COLOMBIA"): (
        "CONVENIO_VENCIDO_COLOMBIA", "ATENCION_CLIENTE",
        "Notificación de cuota vencida convenio Colombia", [],
    ),
    ("CONVENIO", "CONVENIO VENCIDO ARGENTINA"): (
        "CONVENIO_VENCIDO_ARGENTINA", "ATENCION_CLIENTE",
        "Notificación de cuota vencida convenio Argentina", [],
    ),
    ("CONVENIO", "CUOTA VENCIDA ARGENTINA"): (
        "CONVENIO_CUOTA_VENCIDA_AR", "ATENCION_CLIENTE",
        "Cuota vencida convenio Argentina (variante con Coppel)", [],
    ),
    ("OFERTA", "OFERTA PARA REGULARIZACIÓN DE DEUDA"): (
        "OFERTA_REGULARIZACION_DEUDA", "ATENCION_CLIENTE",
        "Oferta personalizada para regularizar deuda", [],
    ),
    ("MSJ AGENCIAS - PEDIDO DOC", "Msj para agencias por pedido doc a banco"): (
        "MSJ_AGENCIAS_PEDIDO_DOC", "ATENCION_CLIENTE",
        "Mensaje a agencias por pedido de documentación al banco", [],
    ),
}

# Filas sin TIPO/MOTIVO (filas 42 y 43 del Excel — identificadas por cuerpo)
# Se mapean por contenido aproximado del cuerpo.
MAPPING_RTAS_SIN_TIPO = [
    ("ADJUNTAMOS_PAZ_Y_SALVO", "ATENCION_CLIENTE",
     "Envío adjunto del paz y salvo solicitado", [],
     "Adjuntamos su paz y salvo"),
    ("CONFIRMACION_RECEPCION_PYS", "ATENCION_CLIENTE",
     "Confirmación automática de recepción de solicitud de PYS", [],
     "su solicitud de paz y salvo fue recibida"),
]

MAPPING_RTA_DC = {
    "Pedido de documentación": (
        "PEDIDO_DOCUMENTACION_DC", "ATENCION_CLIENTE",
        "Datacrédito: solicitar al cliente que escriba al mail corporativo",
        ["enviar documentacion", "enviar documentación", "mandar documentos"],
    ),
    "Desconoce Rappi": (
        "DESCONOCE_DEUDA_RAPPI", "ATENCION_CLIENTE",
        "Cliente desconoce deuda Rappi — explicación compra cartera",
        ["no conozco", "no reconozco", "no es mia", "no es mía"],
        # Necesita además "rappi" para matchear — manejado en _DETECTION_RULES
    ),
    "Desconoce FLB": (
        "DESCONOCE_DEUDA_FALABELLA", "ATENCION_CLIENTE",
        "Cliente desconoce deuda Falabella — explicación compra cartera",
        ["no conozco", "no reconozco", "no es mia"],
    ),
    "Desconoce Santander": (
        "DESCONOCE_DEUDA_SANTANDER", "ATENCION_CLIENTE",
        "Cliente desconoce deuda Santander — explicación compra cartera",
        ["no conozco", "no reconozco", "no es mia"],
    ),
    "Desconoce Bogotá": (
        "DESCONOCE_DEUDA_BANCO_BOGOTA", "ATENCION_CLIENTE",
        "Cliente desconoce deuda Banco Bogotá — explicación compra cartera",
        ["no conozco", "no reconozco", "no es mia"],
    ),
    "Cliente Pide Info": (
        "CLIENTE_PIDE_INFO", "ATENCION_CLIENTE",
        "Cliente pide información — redirigir al mail corporativo",
        ["pido información", "necesito información", "información sobre"],
    ),
    "Cartera Recuperada": (
        "CARTERA_RECUPERADA", "ATENCION_CLIENTE",
        "Confirmar al cliente que su obligación figura como cartera recuperada", [],
    ),
}


def _clean(s) -> str:
    """Normaliza cell de Excel a string sin caracteres raros."""
    if s is None:
        return ""
    return str(s).replace("\xa0", " ").strip()


def cargar_plantillas_del_excel(xlsx_path: str) -> list[dict]:
    """Lee Rtas + RTA DC y devuelve lista de dicts listos para INSERT."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    docs: list[dict] = []

    # ─── Hoja Rtas ─────────────────────────────────────────────────────
    ws = wb["Rtas"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        tipo = _clean(row[1] if len(row) > 1 else "")
        motivo = _clean(row[2] if len(row) > 2 else "")
        mensaje = _clean(row[3] if len(row) > 3 else "")
        if not mensaje:
            continue

        # 1) Match exacto por (TIPO, MOTIVO)
        key = (tipo, motivo)
        if key in MAPPING_RTAS:
            slug, workflow, contexto, keywords = MAPPING_RTAS[key]
            docs.append({
                "slug": slug, "workflow": workflow,
                "contexto": contexto, "keywords": keywords, "cuerpo": mensaje,
                "source_row": f"Rtas:{tipo}/{motivo}",
            })
            continue

        # 2) Sin TIPO/MOTIVO — match por substring del cuerpo
        if not tipo and not motivo:
            for slug, workflow, contexto, keywords, marker in MAPPING_RTAS_SIN_TIPO:
                if marker.lower() in mensaje.lower():
                    docs.append({
                        "slug": slug, "workflow": workflow,
                        "contexto": contexto, "keywords": keywords, "cuerpo": mensaje,
                        "source_row": f"Rtas:NO-TIPO/match={marker[:30]}",
                    })
                    break
            continue

        # 3) Mapping desconocido — log warn y skip
        logger.warning("  ⚠️  Rtas sin mapping: (%r, %r) — skip", tipo, motivo)

    # ─── Hoja RTA DC ───────────────────────────────────────────────────
    ws = wb["RTA DC"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        tipo = _clean(row[0] if len(row) > 0 else "")
        mensaje = _clean(row[1] if len(row) > 1 else "")
        if not mensaje:
            continue
        if tipo in MAPPING_RTA_DC:
            slug, workflow, contexto, keywords = MAPPING_RTA_DC[tipo]
            docs.append({
                "slug": slug, "workflow": workflow,
                "contexto": contexto, "keywords": keywords, "cuerpo": mensaje,
                "source_row": f"RTA DC:{tipo}",
            })
        else:
            logger.warning("  ⚠️  RTA DC sin mapping: %r — skip", tipo)

    return docs


async def upsert(conn: asyncpg.Connection, docs: list[dict], dry_run: bool) -> dict:
    """UPSERT por (cliente_id, problematica). Devuelve contadores."""
    import uuid as _uuid
    contadores = {"insertados": 0, "actualizados": 0}

    for d in docs:
        if dry_run:
            logger.info("  DRY [%s] %s (workflow=%s, chars=%d, kw=%d)",
                        d["source_row"], d["slug"], d["workflow"],
                        len(d["cuerpo"]), len(d["keywords"]))
            continue

        # Existe ya?
        existe = await conn.fetchval(
            "SELECT 1 FROM plantillas_respuesta "
            "WHERE cliente_id = $1::uuid AND problematica = $2",
            TENANT_FF, d["slug"],
        )
        if existe:
            await conn.execute(
                """UPDATE plantillas_respuesta
                   SET cuerpo = $1, contexto = $2, keywords = $3,
                       tipo_workflow = $4, is_active = TRUE
                   WHERE cliente_id = $5::uuid AND problematica = $6""",
                d["cuerpo"], d["contexto"], d["keywords"], d["workflow"],
                TENANT_FF, d["slug"],
            )
            contadores["actualizados"] += 1
        else:
            await conn.execute(
                """INSERT INTO plantillas_respuesta
                      (cliente_id, problematica, contexto, cuerpo, keywords,
                       tipo_workflow, is_active)
                   VALUES ($1::uuid, $2, $3, $4, $5, $6, TRUE)""",
                TENANT_FF, d["slug"], d["contexto"], d["cuerpo"],
                d["keywords"], d["workflow"],
            )
            contadores["insertados"] += 1

    return contadores


async def main_async(args: argparse.Namespace) -> int:
    if not os.path.exists(args.xlsx):
        logger.error("Excel no encontrado: %s", args.xlsx)
        return 2

    logger.info("Leyendo Excel: %s", args.xlsx)
    docs = cargar_plantillas_del_excel(args.xlsx)
    logger.info("Plantillas mapeadas: %d", len(docs))

    if not docs:
        return 0

    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        logger.error("DATABASE_URL no configurada")
        return 2

    conn = await asyncpg.connect(db_url)
    try:
        await conn.execute("SET app.is_superuser = 'true'")
        cont = await upsert(conn, docs, args.dry_run)
        logger.info(
            "Hecho. insertados=%d actualizados=%d (dry_run=%s)",
            cont["insertados"], cont["actualizados"], args.dry_run,
        )
    finally:
        await conn.close()
    return 0


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--xlsx", required=True, help="path a Consolidado Reclamos.xlsx")
    p.add_argument("--dry-run", action="store_true",
                   help="no escribe; solo lista qué upsertaría")
    return p.parse_args()


if __name__ == "__main__":
    sys.exit(asyncio.run(main_async(parse_args())))
