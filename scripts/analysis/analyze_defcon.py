import openpyxl

file_path = "E:\\COLOMBIA\\PQRS_V2\\DEFCON - CO.xlsx"
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
print(f"Hojas: {wb.sheetnames}")

ws = wb["Hoja1"]
rows = list(ws.iter_rows(max_row=5, values_only=True))
for i, row in enumerate(rows):
    print(f"Fila {i}: {row}")
